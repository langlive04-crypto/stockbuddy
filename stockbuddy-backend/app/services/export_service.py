"""
匯出服務 V10.15
支援 CSV / Excel 匯出功能

提供功能：
1. 推薦股票清單匯出
2. 投資組合匯出
3. 交易記錄匯出
4. 回測結果匯出
5. 績效分析報告匯出
"""

import csv
import io
import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """匯出服務"""

    @staticmethod
    def export_recommendations_csv(recommendations: List[Dict]) -> str:
        """
        匯出推薦股票清單為 CSV

        Args:
            recommendations: 推薦股票列表

        Returns:
            CSV 字串
        """
        if not recommendations:
            return ""

        output = io.StringIO()
        writer = csv.writer(output)

        # 標題列
        headers = [
            "股票代號", "股票名稱", "現價", "漲跌幅(%)",
            "AI評分", "訊號", "技術面", "基本面", "籌碼面", "新聞面",
            "本益比", "殖利率(%)", "產業", "分析理由"
        ]
        writer.writerow(headers)

        # 資料列
        for stock in recommendations:
            breakdown = stock.get("score_breakdown", {})
            row = [
                stock.get("stock_id", ""),
                stock.get("name", ""),
                stock.get("price", ""),
                stock.get("change_percent", ""),
                stock.get("confidence", ""),
                stock.get("signal", ""),
                breakdown.get("technical", ""),
                breakdown.get("fundamental", ""),
                breakdown.get("chip", ""),
                breakdown.get("news", 50),
                stock.get("pe_ratio", ""),
                stock.get("dividend_yield", ""),
                stock.get("industry", ""),
                stock.get("reason", ""),
            ]
            writer.writerow(row)

        return output.getvalue()

    @staticmethod
    def export_recommendations_excel(recommendations: List[Dict]) -> bytes:
        """
        匯出推薦股票清單為 Excel

        Args:
            recommendations: 推薦股票列表

        Returns:
            Excel 檔案內容 (bytes)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "AI推薦股票"

        # 樣式定義
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        up_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        down_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 標題列
        headers = [
            "股票代號", "股票名稱", "現價", "漲跌幅(%)",
            "AI評分", "訊號", "技術面", "基本面", "籌碼面", "新聞面",
            "本益比", "殖利率(%)", "產業", "分析理由"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # 資料列
        for row_idx, stock in enumerate(recommendations, 2):
            breakdown = stock.get("score_breakdown", {})
            change_pct = stock.get("change_percent", 0)

            data = [
                stock.get("stock_id", ""),
                stock.get("name", ""),
                stock.get("price", ""),
                change_pct,
                stock.get("confidence", ""),
                stock.get("signal", ""),
                breakdown.get("technical", ""),
                breakdown.get("fundamental", ""),
                breakdown.get("chip", ""),
                breakdown.get("news", 50),
                stock.get("pe_ratio", ""),
                stock.get("dividend_yield", ""),
                stock.get("industry", ""),
                stock.get("reason", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border

                # 漲跌幅顏色
                if col == 4:  # 漲跌幅欄位
                    if change_pct > 0:
                        cell.fill = up_fill
                    elif change_pct < 0:
                        cell.fill = down_fill

        # 調整欄寬
        column_widths = [10, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 儲存到 bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_portfolio_csv(holdings: List[Dict], summary: Dict = None) -> str:
        """
        匯出投資組合為 CSV

        Args:
            holdings: 持股列表
            summary: 摘要統計

        Returns:
            CSV 字串
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # 摘要資訊
        if summary:
            writer.writerow(["===== 投資組合摘要 ====="])
            writer.writerow(["總市值", summary.get("total_value", 0)])
            writer.writerow(["總成本", summary.get("total_cost", 0)])
            writer.writerow(["未實現損益", summary.get("unrealized_pnl", 0)])
            writer.writerow(["報酬率(%)", summary.get("return_pct", 0)])
            writer.writerow([])

        # 持股明細
        writer.writerow(["===== 持股明細 ====="])
        headers = [
            "股票代號", "股票名稱", "持股數量", "買入價格", "現價",
            "市值", "成本", "未實現損益", "報酬率(%)", "買入日期", "備註"
        ]
        writer.writerow(headers)

        for holding in holdings:
            row = [
                holding.get("stock_id", ""),
                holding.get("name", ""),
                holding.get("quantity", 0),
                holding.get("buy_price", 0),
                holding.get("current_price", 0),
                holding.get("market_value", 0),
                holding.get("cost", 0),
                holding.get("unrealized_pnl", 0),
                holding.get("return_pct", 0),
                holding.get("buy_date", ""),
                holding.get("note", ""),
            ]
            writer.writerow(row)

        return output.getvalue()

    @staticmethod
    def export_portfolio_excel(holdings: List[Dict], summary: Dict = None) -> bytes:
        """
        匯出投資組合為 Excel

        Args:
            holdings: 持股列表
            summary: 摘要統計

        Returns:
            Excel 檔案內容 (bytes)
        """
        wb = Workbook()

        # ===== 摘要頁 =====
        ws_summary = wb.active
        ws_summary.title = "投資組合摘要"

        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ws_summary.cell(row=1, column=1, value="投資組合報告").font = Font(bold=True, size=18)
        ws_summary.cell(row=2, column=1, value=f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        if summary:
            ws_summary.cell(row=4, column=1, value="總市值").font = Font(bold=True)
            ws_summary.cell(row=4, column=2, value=summary.get("total_value", 0))

            ws_summary.cell(row=5, column=1, value="總成本").font = Font(bold=True)
            ws_summary.cell(row=5, column=2, value=summary.get("total_cost", 0))

            ws_summary.cell(row=6, column=1, value="未實現損益").font = Font(bold=True)
            ws_summary.cell(row=6, column=2, value=summary.get("unrealized_pnl", 0))

            ws_summary.cell(row=7, column=1, value="報酬率(%)").font = Font(bold=True)
            ws_summary.cell(row=7, column=2, value=summary.get("return_pct", 0))

        # ===== 持股明細頁 =====
        ws_holdings = wb.create_sheet("持股明細")

        headers = [
            "股票代號", "股票名稱", "持股數量", "買入價格", "現價",
            "市值", "成本", "未實現損益", "報酬率(%)", "買入日期", "備註"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_holdings.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, holding in enumerate(holdings, 2):
            data = [
                holding.get("stock_id", ""),
                holding.get("name", ""),
                holding.get("quantity", 0),
                holding.get("buy_price", 0),
                holding.get("current_price", 0),
                holding.get("market_value", 0),
                holding.get("cost", 0),
                holding.get("unrealized_pnl", 0),
                holding.get("return_pct", 0),
                holding.get("buy_date", ""),
                holding.get("note", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = ws_holdings.cell(row=row_idx, column=col, value=value)

                # 損益顏色
                if col == 8:  # 未實現損益
                    if value > 0:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # 調整欄寬
        for col in range(1, 12):
            ws_holdings.column_dimensions[get_column_letter(col)].width = 12

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_backtest_csv(result: Dict) -> str:
        """
        匯出回測結果為 CSV

        Args:
            result: 回測結果

        Returns:
            CSV 字串
        """
        output = io.StringIO()
        writer = csv.writer(output)

        stats = result.get("stats", {})

        # 回測摘要
        writer.writerow(["===== 回測結果摘要 ====="])
        writer.writerow(["股票代號", result.get("stock_id", "")])
        writer.writerow(["策略", result.get("strategy", "")])
        writer.writerow(["期間", result.get("period", "")])
        writer.writerow([])

        # 績效統計
        writer.writerow(["===== 績效統計 ====="])
        writer.writerow(["初始資金", stats.get("initial_capital", 0)])
        writer.writerow(["最終淨值", stats.get("final_value", 0)])
        writer.writerow(["總報酬", stats.get("total_return", 0)])
        writer.writerow(["總報酬率(%)", stats.get("total_return_pct", 0)])
        writer.writerow(["年化報酬率(%)", stats.get("annual_return_pct", 0)])
        writer.writerow(["最大回撤(%)", stats.get("max_drawdown_pct", 0)])
        writer.writerow(["夏普比率", stats.get("sharpe_ratio", 0)])
        writer.writerow(["勝率(%)", stats.get("win_rate", 0)])
        writer.writerow(["獲利因子", stats.get("profit_factor", 0)])
        writer.writerow(["總交易次數", stats.get("total_trades", 0)])
        writer.writerow([])

        # 交易記錄
        trades = result.get("trades", [])
        if trades:
            writer.writerow(["===== 交易記錄 ====="])
            writer.writerow(["日期", "類型", "價格", "數量", "金額", "損益", "原因"])

            for trade in trades:
                row = [
                    trade.get("date", ""),
                    trade.get("type", ""),
                    trade.get("price", 0),
                    trade.get("shares", 0),
                    trade.get("cost", trade.get("proceeds", 0)),
                    trade.get("profit", ""),
                    trade.get("reason", ""),
                ]
                writer.writerow(row)

        return output.getvalue()

    @staticmethod
    def export_backtest_excel(result: Dict) -> bytes:
        """
        匯出回測結果為 Excel

        Args:
            result: 回測結果

        Returns:
            Excel 檔案內容 (bytes)
        """
        wb = Workbook()

        # ===== 摘要頁 =====
        ws_summary = wb.active
        ws_summary.title = "回測摘要"

        stats = result.get("stats", {})

        ws_summary.cell(row=1, column=1, value="回測結果報告").font = Font(bold=True, size=18)
        ws_summary.cell(row=2, column=1, value=f"股票: {result.get('stock_id', '')} | 策略: {result.get('strategy', '')}")
        ws_summary.cell(row=3, column=1, value=f"期間: {result.get('period', '')}")

        summary_data = [
            ("初始資金", stats.get("initial_capital", 0)),
            ("最終淨值", stats.get("final_value", 0)),
            ("總報酬", stats.get("total_return", 0)),
            ("總報酬率(%)", stats.get("total_return_pct", 0)),
            ("年化報酬率(%)", stats.get("annual_return_pct", 0)),
            ("最大回撤(%)", stats.get("max_drawdown_pct", 0)),
            ("夏普比率", stats.get("sharpe_ratio", 0)),
            ("勝率(%)", stats.get("win_rate", 0)),
            ("獲利因子", stats.get("profit_factor", 0)),
            ("總交易次數", stats.get("total_trades", 0)),
        ]

        for row_idx, (label, value) in enumerate(summary_data, 5):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # ===== 交易記錄頁 =====
        trades = result.get("trades", [])
        if trades:
            ws_trades = wb.create_sheet("交易記錄")

            headers = ["日期", "類型", "價格", "數量", "金額", "損益", "損益率(%)", "原因"]
            for col, header in enumerate(headers, 1):
                cell = ws_trades.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row_idx, trade in enumerate(trades, 2):
                data = [
                    trade.get("date", ""),
                    "買入" if trade.get("type") == "buy" else "賣出",
                    trade.get("price", 0),
                    trade.get("shares", 0),
                    trade.get("cost", trade.get("proceeds", 0)),
                    trade.get("profit", ""),
                    trade.get("profit_pct", ""),
                    trade.get("reason", ""),
                ]

                for col, value in enumerate(data, 1):
                    cell = ws_trades.cell(row=row_idx, column=col, value=value)

                    # 買賣顏色
                    if col == 2:
                        if value == "買入":
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # ===== 每日淨值頁 =====
        daily_values = result.get("daily_values", [])
        if daily_values:
            ws_daily = wb.create_sheet("每日淨值")

            headers = ["日期", "淨值", "報酬率(%)"]
            for col, header in enumerate(headers, 1):
                cell = ws_daily.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for row_idx, dv in enumerate(daily_values, 2):
                ws_daily.cell(row=row_idx, column=1, value=dv.get("date", ""))
                ws_daily.cell(row=row_idx, column=2, value=dv.get("value", 0))
                ws_daily.cell(row=row_idx, column=3, value=dv.get("return_pct", 0))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_performance_report(performance: Dict) -> bytes:
        """
        匯出績效分析報告為 Excel

        Args:
            performance: 績效分析結果

        Returns:
            Excel 檔案內容 (bytes)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "績效分析報告"

        header_font = Font(bold=True, size=14)
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 標題
        ws.cell(row=1, column=1, value="績效分析報告").font = Font(bold=True, size=20)
        ws.cell(row=2, column=1, value=f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        current_row = 4

        # 總覽
        summary = performance.get("summary", {})
        ws.cell(row=current_row, column=1, value="績效總覽").font = header_font
        current_row += 1

        summary_items = [
            ("總報酬率", f"{summary.get('total_return_pct', 0):.2f}%"),
            ("年化報酬率", f"{summary.get('annualized_return_pct', 0):.2f}%"),
            ("交易天數", summary.get('trading_days', 0)),
        ]
        for label, value in summary_items:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 1

        # 風險調整指標
        risk_adj = performance.get("risk_adjusted", {})
        ws.cell(row=current_row, column=1, value="風險調整指標").font = header_font
        current_row += 1

        risk_items = [
            ("Alpha", f"{risk_adj.get('alpha', 0):.2f}%"),
            ("Beta", f"{risk_adj.get('beta', 1):.2f}"),
            ("夏普比率", f"{risk_adj.get('sharpe_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{risk_adj.get('sortino_ratio', 0):.2f}"),
            ("Calmar Ratio", f"{risk_adj.get('calmar_ratio', 0):.2f}"),
        ]
        for label, value in risk_items:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 1

        # 風險指標
        risk_metrics = performance.get("risk_metrics", {})
        ws.cell(row=current_row, column=1, value="風險指標").font = header_font
        current_row += 1

        risk_metric_items = [
            ("最大回撤", f"{risk_metrics.get('max_drawdown_pct', 0):.2f}%"),
            ("VaR (95%)", f"{risk_metrics.get('var_95', 0):.2f}%"),
            ("CVaR (95%)", f"{risk_metrics.get('cvar_95', 0):.2f}%"),
            ("年化波動率", f"{risk_metrics.get('volatility_annual', 0):.2f}%"),
        ]
        for label, value in risk_metric_items:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


# 便捷函數
def export_to_csv(data_type: str, data: Any, **kwargs) -> str:
    """匯出為 CSV"""
    if data_type == "recommendations":
        return ExportService.export_recommendations_csv(data)
    elif data_type == "portfolio":
        return ExportService.export_portfolio_csv(data, kwargs.get("summary"))
    elif data_type == "backtest":
        return ExportService.export_backtest_csv(data)
    else:
        return ""

def export_to_excel(data_type: str, data: Any, **kwargs) -> bytes:
    """匯出為 Excel"""
    if data_type == "recommendations":
        return ExportService.export_recommendations_excel(data)
    elif data_type == "portfolio":
        return ExportService.export_portfolio_excel(data, kwargs.get("summary"))
    elif data_type == "backtest":
        return ExportService.export_backtest_excel(data)
    elif data_type == "performance":
        return ExportService.export_performance_report(data)
    else:
        return b""
